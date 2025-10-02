#!/usr/bin/env python3
"""
Excel Report Generator for CT Chest Pathology Classification

Generates Excel reports with study results according to hackathon requirements.
"""

import logging
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple
import uuid

import numpy as np
import SimpleITK as sitk
from minio import Minio
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Generate Excel reports for CT pathology classification"""

    def __init__(self):
        self.results: List[Dict] = []

    def add_study_result(
        self,
        path_to_study: str,
        study_uid: str,
        series_uid: str,
        probability_of_pathology: float,
        pathology: int,
        processing_status: str,
        time_of_processing: float,
        pathology_localization: str,
    ):
        """
        Add a study result to the report

        Args:
            path_to_study: Path to DICOM study ZIP file
            study_uid: StudyInstanceUID from DICOM
            series_uid: SeriesInstanceUID from DICOM
            probability_of_pathology: Pathology probability (0.0 or 1.0)
            pathology: Binary pathology indicator (0 or 1)
            processing_status: "Success" or "Failure"
            time_of_processing: Processing time in seconds
            pathology_localization: Bounding box string "x_min,x_max,y_min,y_max,z_min,z_max"
        """
        self.results.append(
            {
                "path_to_study": path_to_study,
                "study_uid": study_uid,
                "series_uid": series_uid,
                "probability_of_pathology": probability_of_pathology,
                "pathology": pathology,
                "processing_status": processing_status,
                "time_of_processing": time_of_processing,
                "pathology_localization": pathology_localization,
            }
        )

    def generate_report(self, output_dir: Path) -> Path:
        """
        Generate Excel report with all study results

        Args:
            output_dir: Directory to save the report

        Returns:
            Path to generated Excel file
        """
        output_dir.mkdir(parents=True, exist_ok=True)

        # Generate unique filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_id = str(uuid.uuid4())[:8]
        filename = f"report_{timestamp}_{report_id}.xlsx"
        output_path = output_dir / filename

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "CT Pathology Classification"

        # Headers
        headers = [
            "path_to_study",
            "study_uid",
            "series_uid",
            "probability_of_pathology",
            "pathology",
            "processing_status",
            "time_of_processing",
            "pathology_localization",
        ]

        # Style headers
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        for row_num, result in enumerate(self.results, 2):
            ws.cell(row=row_num, column=1, value=result["path_to_study"])
            ws.cell(row=row_num, column=2, value=result["study_uid"])
            ws.cell(row=row_num, column=3, value=result["series_uid"])
            ws.cell(row=row_num, column=4, value=result["probability_of_pathology"])
            ws.cell(row=row_num, column=5, value=result["pathology"])
            ws.cell(row=row_num, column=6, value=result["processing_status"])
            ws.cell(row=row_num, column=7, value=result["time_of_processing"])
            ws.cell(row=row_num, column=8, value=result["pathology_localization"])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(output_path)

        logger.info(f"Generated Excel report: {output_path}")
        return output_path


def determine_pathology(metrics: Dict) -> Tuple[float, int]:
    """
    Determine pathology presence from segmentation metrics

    Args:
        metrics: Dictionary with 'detected_classes' key containing segmentation results

    Returns:
        (probability_of_pathology, pathology_binary)
        - probability: 1.0 if pathology detected, 0.0 otherwise
        - pathology_binary: 1 if pathology detected, 0 otherwise
    """
    detected_classes = metrics.get("detected_classes", {})

    # Check if any class exists (background is excluded in detected_classes)
    # If detected_classes is not empty, pathology is present
    has_pathology = len(detected_classes) > 0

    if has_pathology:
        logger.info(f"Pathology detected: {list(detected_classes.keys())}")
        return (1.0, 1)
    else:
        logger.info("No pathology detected (normal study)")
        return (0.0, 0)


def calculate_pathology_bbox(
    minio_path: str, minio_client: Minio, bucket_name: str
) -> str:
    """
    Calculate 3D bounding box of pathology from segmentation mask

    Args:
        minio_path: Path to segmentation file in MinIO
        minio_client: MinIO client instance
        bucket_name: MinIO bucket name

    Returns:
        Comma-separated bounding box: "x_min,x_max,y_min,y_max,z_min,z_max"
        Empty string if no pathology found
    """
    try:
        # Download segmentation from MinIO to temp file
        with tempfile.NamedTemporaryFile(suffix=".nii.gz", delete=False) as tmp:
            tmp_path = tmp.name

        minio_client.fget_object(bucket_name, minio_path, tmp_path)

        # Load NIfTI
        seg_image = sitk.ReadImage(tmp_path)
        seg_array = sitk.GetArrayFromImage(seg_image)  # Shape: (Z, Y, X)

        # Find all non-zero voxels (pathology)
        pathology_mask = seg_array > 0
        nonzero_indices = np.where(pathology_mask)

        if len(nonzero_indices[0]) == 0:
            logger.warning("No pathology voxels found in segmentation")
            return ""

        # Calculate bounding box
        # nonzero_indices[0] = Z coordinates
        # nonzero_indices[1] = Y coordinates
        # nonzero_indices[2] = X coordinates
        z_min, z_max = int(nonzero_indices[0].min()), int(nonzero_indices[0].max())
        y_min, y_max = int(nonzero_indices[1].min()), int(nonzero_indices[1].max())
        x_min, x_max = int(nonzero_indices[2].min()), int(nonzero_indices[2].max())

        # Format as string: x_min,x_max,y_min,y_max,z_min,z_max
        bbox_str = f"{x_min},{x_max},{y_min},{y_max},{z_min},{z_max}"

        logger.info(f"Calculated pathology bounding box: {bbox_str}")
        return bbox_str

    except Exception as e:
        logger.error(f"Error calculating bounding box: {e}")
        return ""
    finally:
        # Cleanup temp file
        try:
            Path(tmp_path).unlink()
        except Exception:
            pass


# Example usage
if __name__ == "__main__":
    # Configure logging
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
    )

    # Example: Create report with sample data
    generator = ExcelReportGenerator()

    generator.add_study_result(
        path_to_study="data/studies/patient_001.zip",
        study_uid="1.2.840.113619.2.55.3.12345678",
        series_uid="1.2.840.113619.2.55.3.87654321",
        probability_of_pathology=1.0,
        pathology=1,
        processing_status="Success",
        time_of_processing=45.23,
        pathology_localization="120,350,80,420,50,180",
    )

    generator.add_study_result(
        path_to_study="data/studies/patient_002.zip",
        study_uid="1.2.840.113619.2.55.3.11111111",
        series_uid="1.2.840.113619.2.55.3.22222222",
        probability_of_pathology=0.0,
        pathology=0,
        processing_status="Success",
        time_of_processing=38.91,
        pathology_localization="",
    )

    report_path = generator.generate_report(output_dir=Path("data/reports"))
    print(f"✓ Report generated: {report_path}")
